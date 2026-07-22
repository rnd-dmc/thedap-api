from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

from THEDAP_REPORT.DapReportStyler import *

def DapReportReachCurve(reportOption, reportCurve, target_pop):
    
    wb = Workbook()

    ### 미디어믹스
    ws = wb.active
    ws.title = "미디어믹스"

    ws.sheet_view.zoomScale = 90
    ws.column_dimensions['A'].width = 3.86
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False

    mix_xl = pd.DataFrame(reportOption.get("input_mix"))
    sum_budget = np.sum(mix_xl['budget'].fillna(.0))

    ##
    ws.cell(row=2, column=2, value="분석일자").style = index_style
    ws.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws.cell(row=3, column=2, value="총 예산").style = index_style
    ws.cell(row=3, column=3, value=sum_budget).style = title_style

    ws.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws.cell(row=4, column=3, value=f"{reportOption['input_gender']}{reportOption['input_age_min']}{reportOption['input_age_max']}").style = title_style

    ws.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws.cell(row=5, column=3, value=target_pop).style = title_style

    ##
    ws.cell(row=7, column=2, value="미디어믹스").style = index_style

    colname_dict = {
        'campaign':'캠페인', 'platform':'매체', 'product':'광고상품', 'date_start':'시작일', 'date_end':'종료일', 'gender':'성별',
        'min':'min', 'max':'max', 'retargeting':'리타겟팅 모수', 'impact':'Impact 가중치', 
        'budget':'예산', 'bid_type':'단가유형', 'bid_cost':'단가', 'bid_rate':'효율', 'imp':'IMP / GRP', 'reach':'Reach'
    }
    for col_idx, col_name in enumerate(mix_xl.columns, start=2):
        cell = ws.cell(row=9, column=col_idx, value=colname_dict[col_name])
        cell.style = index_style

    ## 
    for row_idx, row in mix_xl.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=10 + row_idx, column=col_idx, value=val)
            if col_idx in [2, 3, 4, 5, 6, 7, 8, 9, 13]:
                cell.style = text_style
            elif col_idx in [10, 12, 14, 16, 17]:
                cell.style = integer_style
            elif col_idx in [11]:
                cell.style = percent_style_round
            else:
                cell.style = percent_style

    for letter_ in [get_column_letter(i) for i in range(2, 18)]:
        if letter_ in ['B', 'C', 'J']:
            ws.column_dimensions[letter_].width = 20
        elif letter_ in ['D']:
            ws.column_dimensions[letter_].width = 30
        elif letter_ in ['G', 'H', 'I', 'O']:
            ws.column_dimensions[letter_].width = 7
        else :
            ws.column_dimensions[letter_].width = 15
            
    ### 분석결과
    ws2 = wb.create_sheet("분석결과")
    ws2.sheet_view.zoomScale = 90
    ws2.column_dimensions['A'].width = 3.86
    ws2.row_dimensions[1].height = 24
    ws2.sheet_view.showGridLines = False

    ##
    ws2.cell(row=2, column=2, value="분석일자").style = index_style
    ws2.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws2.cell(row=3, column=2, value="총 예산").style = index_style
    ws2.cell(row=3, column=3, value=sum_budget).style = title_style

    ws2.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws2.cell(row=4, column=3, value=f"{reportOption['input_gender']}{reportOption['input_age_min']}{reportOption['input_age_max']}").style = title_style

    ws2.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws2.cell(row=5, column=3, value=target_pop).style = title_style

    ##
    ws2.cell(row=7, column=2, value="리치커브").style = index_style

    curve_xl = pd.DataFrame(reportCurve).drop('idx', axis=1).query('budget > .0').reset_index(drop=True).filter(regex=r'(?<!_n)$')
    colname_dict2 = {
        'line':'구분', 'campaign':'캠페인', 'platform':'매체', 'product':'광고상품', 'gender':'성별', 'age_min':'Min', 'age_max':'Max',
        'budget':'예산', 'date_start':'시작일', 'date_end':'종료일', 
        'target_impression':'Target Impression', 'target_impression_weighted':'Target Impression (가중치+)', 'target_grps':'Target GRPs', 'target_grps_weighted':'Target GRPs (가중치+)',
        'target_reach_n':'Target Reach(n)', 'target_reach2_n':'Target Reach 2+(n)', 'target_reach3_n':'Target Reach 3+(n)', 'target_reach4_n':'Target Reach 4+(n)', 'target_reach5_n':'Target Reach 5+(n)', 
        'target_reach6_n':'Target Reach 6+(n)', 'target_reach7_n':'Target Reach 7+(n)', 'target_reach8_n':'Target Reach 8+(n)', 'target_reach9_n':'Target Reach 9+(n)', 'target_reach10_n':'Target Reach 10+(n)', 
        'target_reach_p':'Target Reach(%)', 'target_reach2_p':'Target Reach 2+(%)', 'target_reach3_p':'Target Reach 3+(%)', 'target_reach4_p':'Target Reach 4+(%)', 'target_reach5_p':'Target Reach 5+(%)',
        'target_reach6_p':'Target Reach 6+(%)', 'target_reach7_p':'Target Reach 7+(%)', 'target_reach8_p':'Target Reach 8+(%)', 'target_reach9_p':'Target Reach 9+(%)', 'target_reach10_p':'Target Reach 10+(%)', 
        'target_af':'Target A.F', 'target_af_day':'Target A.F (1 Day)', 'target_af_week':'Target A.F (Week)', 'target_af_30':'Target A.F (30 Days)', 'imp_interval':'Impression Interval'
    }
    for letter_ in [get_column_letter(i) for i in range(2, 16)]:
        if letter_ in ['B', 'C', 'J']:
            ws2.column_dimensions[letter_].width = 20
        elif letter_ in ['O']:
            ws2.column_dimensions[letter_].width = 12
        else:
            ws2.column_dimensions[letter_].width = 18
        
    for col_idx, col_name in enumerate(curve_xl.columns, start=2):
        cell = ws2.cell(row=9, column=col_idx, value=colname_dict2[col_name])
        cell.style = index_style

    for row_idx, row in curve_xl.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
            if col_idx in [2]:
                cell.style = integer_style
            elif col_idx in [3, 4, 15]:
                cell.style = float_style
            elif col_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                cell.style = percent_style
                
    for row_idx, row in curve_xl.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)

    # Databat Rule
    databar_rule = DataBarRule(
        start_type='num', start_value=0,
        end_type='num', end_value=1,
        color="6D8AEF",  
        showValue=True,
        minLength=0, maxLength=100
    )

    for col in range(5, 14+1):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{10}:{col_letter}{10+curve_xl.shape[0]}"
        ws2.conditional_formatting.add(cell_range, databar_rule)
        
    return wb