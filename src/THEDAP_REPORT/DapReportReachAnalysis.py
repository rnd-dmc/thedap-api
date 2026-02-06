from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

from THEDAP_REPORT.DapReportStyler import *

def DapReportReachAnalysis(reportOption, reportResult, target_pop, userGrade):
    
    wb = Workbook()
    ### 미디어믹스
    ws = wb.active
    ws.title = "미디어믹스"

    ws.sheet_view.zoomScale = 90
    ws.column_dimensions['A'].width = 3.86
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False

    mix_xl = pd.DataFrame(reportOption.get("input_mix"))
    sum_budget = np.sum(mix_xl['budget'].fillna(.0))

    ##
    ws.cell(row=2, column=2, value="분석일자").style = index_style
    ws.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws.cell(row=3, column=2, value="총 예산").style = index_style
    ws.cell(row=3, column=3, value=sum_budget).style = title_style

    ws.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws.cell(row=4, column=3, value=f"{reportOption['input_gender']}{reportOption['input_age_min']}{reportOption['input_age_max']}").style = title_style

    ws.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws.cell(row=5, column=3, value=target_pop).style = title_style

    ws.cell(row=6, column=2, value="데이터 기준").style = index_style
    ws.cell(row=6, column=3, value=reportOption['inputModelDate']).style = title_style
    
    ##
    ws.cell(row=8, column=2, value="미디어믹스").style = index_style
    
    ### 분석결과
    ws2 = wb.create_sheet("분석결과")
    ws2.sheet_view.zoomScale = 90
    ws2.column_dimensions['A'].width = 3.86
    ws2.row_dimensions[1].height = 24
    ws2.sheet_view.showGridLines = False
    
    ws2.cell(row=2, column=2, value="분석일자").style = index_style
    ws2.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws2.cell(row=3, column=2, value="총 예산").style = index_style
    ws2.cell(row=3, column=3, value=sum_budget).style = title_style

    ws2.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws2.cell(row=4, column=3, value=f"{reportOption['input_gender']}{reportOption['input_age_min']}{reportOption['input_age_max']}").style = title_style

    ws2.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws2.cell(row=5, column=3, value=target_pop).style = title_style

    ws2.cell(row=6, column=2, value="데이터 기준").style = index_style
    ws2.cell(row=6, column=3, value=reportOption['inputModelDate']).style = title_style
    
    if userGrade != "B":

        colname_dict = {
            'campaign':'캠페인', 'platform':'광고 매체', 'product':'광고 상품', 'date_start':'시작일', 'date_end':'종료일', 'gender':'성별',
            'min':'min', 'max':'max', 'retargeting':'타겟팅 모수', 'impact':'Impact 가중치', 
            'budget':'예산', 'bid_type':'단가유형', 'bid_cost':'단가', 'bid_rate':'효율', 'imp':'IMP / GRP', 'reach':'Reach'
        }
        for col_idx, col_name in enumerate(mix_xl.columns, start=2):
            cell = ws.cell(row=9, column=col_idx, value=colname_dict[col_name])
            cell.style = index_style

        ## 
        for row_idx, row in mix_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell = ws.cell(row=10 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3, 4, 5, 6, 7, 8, 9, 13]:
                    cell.style = text_style
                elif col_idx in [10, 12, 14, 16, 17]:
                    cell.style = integer_style
                elif col_idx in [11]:
                    cell.style = percent_style_round
                else:
                    cell.style = percent_style

        for letter_ in [get_column_letter(i) for i in range(2, 18)]:
            if letter_ in ['B', 'C', 'J']:
                ws.column_dimensions[letter_].width = 20
            elif letter_ in ['D']:
                ws.column_dimensions[letter_].width = 30
            elif letter_ in ['G', 'H', 'I', 'O']:
                ws.column_dimensions[letter_].width = 7
            else :
                ws.column_dimensions[letter_].width = 15

        ### 분석결과        
        ##
        ws2.cell(row=8, column=2, value="분석결과").style = index_style
        colname_dict2 = {
            'line':'구분', 'campaign':'캠페인', 'platform':'광고 매체', 'product':'광고 상품', 'gender':'성별', 'age_min':'Min', 'age_max':'Max',
            'budget':'예산', 'date_start':'시작일', 'date_end':'종료일', 
            'target_impression':'Target Impression', 'target_impression_weighted':'Target Impression (가중치+)', 'target_grps':'Target GRPs', 'target_grps_weighted':'Target GRPs (가중치+)',
            'target_reach_n':'Target Reach(n)', 'target_reach2_n':'Target Reach 2+(n)', 'target_reach3_n':'Target Reach 3+(n)', 'target_reach4_n':'Target Reach 4+(n)', 'target_reach5_n':'Target Reach 5+(n)', 
            'target_reach6_n':'Target Reach 6+(n)', 'target_reach7_n':'Target Reach 7+(n)', 'target_reach8_n':'Target Reach 8+(n)', 'target_reach9_n':'Target Reach 9+(n)', 'target_reach10_n':'Target Reach 10+(n)', 
            'target_reach_p':'Target Reach(%)', 'target_reach2_p':'Target Reach 2+(%)', 'target_reach3_p':'Target Reach 3+(%)', 'target_reach4_p':'Target Reach 4+(%)', 'target_reach5_p':'Target Reach 5+(%)',
            'target_reach6_p':'Target Reach 6+(%)', 'target_reach7_p':'Target Reach 7+(%)', 'target_reach8_p':'Target Reach 8+(%)', 'target_reach9_p':'Target Reach 9+(%)', 'target_reach10_p':'Target Reach 10+(%)', 
            'target_af':'Target A.F', 'target_af_weighted':'Target A.F (가중치+)', 'target_af_day':'Target A.F (1 Day)', 'target_af_week':'Target A.F (Week)', 'target_af_30':'Target A.F (30 Days)', 'imp_interval':'Impression Interval'
        }

        summary_xl = pd.DataFrame(reportResult['result_summary']).\
            filter(
                [
                    'line', 'campaign', 'platform', 'product', 'gender', 'age_min', 'age_max', 'budget', 'date_start', 'date_end', 'target_impression', 'target_grps', 
                    'target_reach_n', 'target_reach_p', 'target_reach2_n', 'target_reach2_p', 'target_reach3_n', 'target_reach3_p', 'target_reach4_n', 'target_reach4_p', 'target_reach5_n', 'target_reach5_p',
                    'target_reach6_n', 'target_reach6_p', 'target_reach7_n', 'target_reach7_p', 'target_reach8_n', 'target_reach8_p', 'target_reach9_n', 'target_reach9_p', 'target_reach10_n', 'target_reach10_p',
                    'target_impression_weighted', 'target_grps_weighted', 'target_af', 'target_af_weighted', 'target_af_day', 'target_af_week', 'target_af_30', 'imp_interval'
                ]
            )

        for letter_ in [get_column_letter(i) for i in range(2, 42)]:
            if letter_ in ['B', 'C', 'D']:
                ws2.column_dimensions[letter_].width = 20
            elif letter_ in ['E']:
                ws2.column_dimensions[letter_].width = 30
            elif letter_ in ['F', 'G', 'H']:
                ws2.column_dimensions[letter_].width = 4
            else:
                ws2.column_dimensions[letter_].width = 18

        subttls = list(np.where(summary_xl['line'].apply(lambda x: x == 'Sub Total'))[0])
        camttls = list(np.where(summary_xl['line'].apply(lambda x: x == 'Campaign Total'))[0])
        plttls = list(np.where(summary_xl['line'].apply(lambda x: x == 'Platform Total'))[0])
        ttl = list(np.where(summary_xl['line'].apply(lambda x: x == 'Total'))[0])

        for col_idx, col_name in enumerate(summary_xl.columns, start=2):
            cell = ws2.cell(row=9, column=col_idx, value=colname_dict2[col_name])
            cell.style = index_style

        for row_idx, row in summary_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3, 4, 5, 6, 10, 11, 41]:
                    cell.style = text_style
                elif col_idx in [7, 8, 9, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34]:
                    cell.style = integer_style
                elif col_idx in [13, 40, 39, 38, 37, 36, 35]:
                    cell.style = float_style     
                else:
                    cell.style = percent_style

        for row_idx, row in summary_xl.iterrows():
            if row_idx in subttls:
                ws2.merge_cells(start_row=10+row_idx, start_column=4, end_row=10+row_idx, end_column=8)
                
            elif row_idx in camttls:
                for col_idx, val in enumerate(row, start=2):
                    cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                ws2.merge_cells(start_row=10+row_idx, start_column=3, end_row=10+row_idx, end_column=8)

            elif row_idx in plttls:
                for col_idx, val in enumerate(row, start=2):
                    cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
                ws2.merge_cells(start_row=10+row_idx, start_column=4, end_row=10+row_idx, end_column=8)
                ws2.merge_cells(start_row=10+row_idx, start_column=2, end_row=10+row_idx, end_column=3)
                
            elif row_idx in ttl:
                for col_idx, val in enumerate(row, start=2):
                    cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                    cell.fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                ws2.merge_cells(start_row=10+row_idx, start_column=2, end_row=10+row_idx, end_column=8)

        ws2.auto_filter.ref = "B9:B9"
    else:
        colname_dict = {
            'platform':'광고 매체', 'product':'광고 상품', 'gender':'성별',
            'min':'min', 'max':'max', 'impact':'Impact 가중치', 
            'budget':'예산', 'bid_type':'단가유형', 'bid_cost':'단가', 'bid_rate':'효율', 'e_imp':'E.IMP', 'e_grp':'E.GRP'
        }
        for col_idx, col_name in enumerate(mix_xl.columns, start=2):
            cell = ws.cell(row=9, column=col_idx, value=colname_dict[col_name])
            cell.style = index_style

        ## 
        for row_idx, row in mix_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell = ws.cell(row=10 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3, 4, 5, 6, 9]:
                    cell.style = text_style
                elif col_idx in [8, 10, 11, 12, 13]:
                    cell.style = integer_style
                elif col_idx in [7]:
                    cell.style = percent_style_round
                else:
                    cell.style = percent_style
                
        for letter_ in [get_column_letter(i) for i in range(2, 14)]:
            if letter_ in ['B', 'C']:
                ws.column_dimensions[letter_].width = 20
            elif letter_ in ['D', 'E', 'F', 'K']:
                ws.column_dimensions[letter_].width = 7
            else:
                ws.column_dimensions[letter_].width = 15
        
        ws2.cell(row=8, column=2, value="분석결과").style = index_style
        colname_dict2 = {
            'line':'구분', 'platform':'광고 매체', 'product':'광고 상품', 'gender':'성별', 'age_min':'Min', 'age_max':'Max',
            'budget':'예산', 
            'target_impression':'Target Impression', 'target_impression_weighted':'Target Impression (가중치+)', 'target_grps':'Target GRPs', 'target_grps_weighted':'Target GRPs (가중치+)',
            'target_reach_n':'Target Reach(n)', 
            'target_reach_p':'Target Reach(%)', 'target_af':'Target A.F'
        }
        
        summary_xl = pd.DataFrame(reportResult['result_summary']).\
            filter(
                [
                    'line', 'platform', 'product', 'gender', 'age_min', 'age_max', 'budget', 'target_impression', 'target_grps', 
                    'target_reach_n', 'target_reach_p', 'target_impression_weighted', 'target_grps_weighted', 'target_af'
                ]
            )
            
        for letter_ in [get_column_letter(i) for i in range(2, 42)]:
            if letter_ in ['B', 'C', 'D']:
                ws2.column_dimensions[letter_].width = 20
            elif letter_ in ['E', 'F', 'G']:
                ws2.column_dimensions[letter_].width = 4
            else:
                ws2.column_dimensions[letter_].width = 18
                
        camttls = list(np.where(summary_xl['line'].apply(lambda x: x == 'Total'))[0])
        ttl = list(np.where(summary_xl['line'].apply(lambda x: x == 'Grand Total'))[0])
        
        for col_idx, col_name in enumerate(summary_xl.columns, start=2):
            cell = ws2.cell(row=9, column=col_idx, value=colname_dict2[col_name])
            cell.style = index_style
            
        for row_idx, row in summary_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3, 4, 5]:
                    cell.style = text_style
                elif col_idx in [6, 7, 8, 9, 11, 13]:
                    cell.style = integer_style
                elif col_idx in [10, 14, 15]:
                    cell.style = float_style     
                else:
                    cell.style = percent_style

        for row_idx, row in summary_xl.iterrows():
            if row_idx in camttls:
                for col_idx, val in enumerate(row, start=2):
                    cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                    cell.fill = PatternFill(start_color="D7DBDD", end_color="D7DBDD", fill_type="solid")
                ws2.merge_cells(start_row=10+row_idx, start_column=3, end_row=10+row_idx, end_column=8)
    
            elif row_idx in ttl:
                for col_idx, val in enumerate(row, start=2):
                    cell = ws2.cell(row=10 + row_idx, column=col_idx, value=val)
                    cell.fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                ws2.merge_cells(start_row=10+row_idx, start_column=2, end_row=10+row_idx, end_column=8)

        ws2.auto_filter.ref = "B9:B9"
        
    target_row = 9 + summary_xl.shape[0] + 1
    start_col = 2
    end_col = 1 + summary_xl.shape[1]

    for col in range(start_col, end_col + 1):
        cell = ws2.cell(row=target_row, column=col)
        cell.border = bottom_border

    ## 히트맵 & 빈도분포
    base_row = 9 + summary_xl.shape[0]  # nrow(rc_tbl_xl)
    row1 = base_row + 3
    row2 = base_row + 4

    # 빈도분포
    ws2.cell(row=row1, column=2, value="도달 히트맵")
    ws2.merge_cells(start_row=row1, end_row=row2, start_column=2, end_column=2)

    ws2.cell(row=row1, column=3, value="히트맵 기준 ↓")
    ws2.cell(row=row2, column=3, value="미디어믹스 전체기준")

    ws2.cell(row=row1, column=4, value="도달 지표 ↓ ")
    ws2.cell(row=row2, column=4, value="Reach(%)")

    ws2.cell(row=row1, column=9, value="도달 빈도")
    ws2.merge_cells(start_row=row1, end_row=row2, start_column=9, end_column=9)

    base_row = 9 + summary_xl.shape[0]
    row1, row2 = base_row + 3, base_row + 4

    ttof = [2, 9]       
    tton = [3, 4]     

    for col in ttof:
        for row in range(row1, row2 + 1):
            cell = ws2.cell(row=row, column=col)
            for k, v in gray_center_bold.items():
                setattr(cell, k, v)

    for idx, col in enumerate(tton):
        cell_top = ws2.cell(row=row1, column=col)
        cell_bot = ws2.cell(row=row2, column=col)

        for k, v in gray_center_bold_small.items():
            setattr(cell_top, k, v)

        for k, v in white_center_normal.items():
            setattr(cell_bot, k, v)
            
    freq_xl = pd.DataFrame(reportResult['reach_freq'])
    freq_xl['reach'] = freq_xl['reach'].apply(lambda x: f'{x}+(%)')
    colname_dict3 = {'reach':'도달 빈도 구분', 'reach_p':'N+ (%)', 'reach_p_diff':'N+ Diff(%)'}

    for col_idx, col_name in enumerate(freq_xl.columns, start=9):
        cell = ws2.cell(row=row2+2, column=col_idx, value=colname_dict3[col_name])
        cell.style = index_style

    for row_idx, row in freq_xl.iterrows():
        for col_idx, val in enumerate(row, start=9):
            cell = ws2.cell(row=row2+3 + row_idx, column=col_idx, value=val)
            cell.style = percent_style2
        
    # 히트맵
    heatmap_xl = pd.DataFrame({
        '연령대':[
            '07-12', '13-18', '19-24', '25-29', '30-34', '35-39', '40-44',
            '45-49', '50-54', '55-59', '60-64', '65-69', '70-74', '75-79'
        ],
        '여성':[None for i in range(14)],
        '남성':[None for i in range(14)]
    })

    for col_idx, col_name in enumerate(heatmap_xl.columns, start=2):
        cell = ws2.cell(row=row2+2, column=col_idx, value=col_name)
        cell.style = index_style
        
    for row_idx, row in heatmap_xl.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws2.cell(row=row2+3 + row_idx, column=col_idx, value=val)
            cell.style = numeric_style

    ### RAW DATA
    ws3 = wb.create_sheet("RAW_DATA")
    ws3.sheet_view.zoomScale = 90
    ws3.column_dimensions['A'].width = 3.86
    ws3.row_dimensions[1].height = 24
    ws3.sheet_view.showGridLines = False

    ws3.column_dimensions["Z"].hidden = True
    ws3.column_dimensions["AA"].hidden = True
    ws3.sheet_state = 'veryHidden'

    heatmap_raw = []

    for k in [j.keys() for j in reportResult['heatmap']][0]:
        stand_ = reportResult['heatmap'][0][k]
        
        for mi, mm in enumerate([list(m.keys())[0] for m in stand_]):
            heatmap_raw.append(pd.DataFrame(stand_[mi][mm]).assign(stand = k, metric = mm))
            
    heatmap_df = pd.concat(heatmap_raw, ignore_index=True)

    heatmap_df_f = heatmap_df.pivot(index=['stand', 'age'], columns='metric', values='F').reset_index().assign(gender='여성')
    heatmap_df_m = heatmap_df.pivot(index=['stand', 'age'], columns='metric', values='M').reset_index().assign(gender='남성')

    heatmap_pv = pd.concat([heatmap_df_f, heatmap_df_m]).reset_index(drop=True)
    heatmap_pv = heatmap_pv.rename(columns={cols:colname_dict2[cols.replace('e_', 'target_')].replace('Target ', '') for cols in heatmap_pv.select_dtypes('float')}).\
        filter(['age', 'gender', 'stand',
                'GRPs','Reach(%)','Reach(n)','Reach 2+(%)','Reach 2+(n)','Reach 3+(%)','Reach 3+(n)','Reach 4+(%)','Reach 4+(n)','Reach 5+(%)','Reach 5+(n)',
                'Reach 6+(%)','Reach 6+(n)','Reach 7+(%)','Reach 7+(n)','Reach 8+(%)','Reach 8+(n)','Reach 9+(%)','Reach 9+(n)','Reach 10+(%)','Reach 10+(n)']).\
        sort_values(['stand', 'gender', 'age']).reset_index(drop=True)
        
    for col_idx, col_name in enumerate(heatmap_pv.columns, start=2):
        cell = ws3.cell(row=2, column=col_idx, value=col_name)
        cell.style = index_style

    for row_idx, row in heatmap_pv.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws3.cell(row=3 + row_idx, column=col_idx, value=val)
            cell.style = numeric_style

    # Validator
    stand_unique = (list(heatmap_pv['stand'].unique()))
    for ii, mm in enumerate(stand_unique):
        ws3.cell(row=3+ii, column=27, value=mm)

    stand_validator = DataValidation(
        type="list", formula1=f"=RAW_DATA!AA$3:AA${3+len(stand_unique)-1}",
        allow_blank=True, showErrorMessage=True, showInputMessage=True
    )

    ws2.add_data_validation(stand_validator)
    stand_validator.add(ws2.cell(row=row2, column=3).coordinate)

    metric_validator = DataValidation(
        type="list", formula1=f"= RAW_DATA!E2:G2",
        allow_blank=True, showErrorMessage=True, showInputMessage=True
    )
    ws2.add_data_validation(metric_validator)
    metric_validator.add(ws2.cell(row=row2, column=4).coordinate)

    ### 
    ws3.cell(row=1, column=27, value=f"=MATCH(분석결과!D{row2}, RAW_DATA!B2:Y2,0)")
    ws3.cell(row=1, column=26, value=f'=IF(RIGHT(분석결과!D{row2}, 3) = "(%)", 1, IF(RIGHT(분석결과!D{row2}, 3) = "(n)", 2500000, MAX(분석결과!C{row2+3}:D{row2+16}) * 1.1))')

    for r in range(heatmap_pv.shape[0]):
        ws3.cell(row=r+3, column=26, value=f'=INDEX(B{r+3}:Y{r+3}, $AA$1)')

    for row_ in range(14):    
        for col_ in range(2):
            ws2.cell(row=row2+3+row_, column=col_+3, 
                    value=f'=SUMIFS(RAW_DATA!Z:Z, RAW_DATA!D:D, 분석결과!$C${row2}, RAW_DATA!C:C, 분석결과!${"C"if col_ == 0 else "D"}${row2+2}, RAW_DATA!B:B, 분석결과!B{row2+3+row_})'
            )
            
    # ColorScale
    color_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color="FFFFFF",
        end_type='formula', end_value="RAW_DATA!$Z$1", end_color="6D8AEF"
    )
    ws2.conditional_formatting.add(f"C{row2+3}:C{row2+3+13}", color_rule)
    ws2.conditional_formatting.add(f"D{row2+3}:D{row2+3+13}", color_rule)
        
    return wb