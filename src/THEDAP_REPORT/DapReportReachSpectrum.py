from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd

from THEDAP_REPORT.DapReportStyler import *

def DapReportReachSpectrum(reportOption, reportOptimize, target_pop):

    wb = Workbook()

    ws = wb.active
    ws.title = "분석결과"

    ws.sheet_view.zoomScale = 90
    ws.column_dimensions['A'].width = 3.86
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False

    maxbudget = reportOption['opt_maxbudget']*1_000_000
    ##
    ws.cell(row=2, column=2, value="분석일자").style = index_style
    ws.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws.cell(row=3, column=2, value="총 예산").style = index_style
    ws.cell(row=3, column=3, value=maxbudget).style = title_style

    ws.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws.cell(row=4, column=3, value=f"{reportOption['input_gender']}{reportOption['input_age_min']}{reportOption['input_age_max']}").style = title_style

    ws.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws.cell(row=5, column=3, value=target_pop).style = title_style
    
    ws.cell(row=6, column=2, value="분석 모델 버전").style = index_style
    ws.cell(row=6, column=3, value=reportOption['inputModelDate']).style = title_style
        
    ##
    colname_dict = {
        'campaign':'캠페인', 'platform':'광고 매체', 'product':'광고 상품', 'date_start':'시작일', 'date_end':'종료일', 'gender':'성별',
        'min':'min', 'max':'max', 'retargeting':'리타겟팅 모수', 'impact':'Impact 가중치', 
        'budget':'예산', 'bid_type':'단가유형', 'bid_cost':'단가', 'bid_rate':'효율', 'imp':'IMP / GRP', 'reach':'Reach', 'alloc_rat':'예산 비중(%)'
    }

    mix_xl_a = pd.DataFrame(reportOption['input_mixA'])
    mix_xl_b = pd.DataFrame(reportOption['input_mixB'])

    ws.cell(row=8, column=2, value="미디어믹스 A").style = index_style

    for col_idx, col_name in enumerate(mix_xl_a.columns, start=2):
        cell = ws.cell(row=9, column=col_idx, value=colname_dict[col_name])
        cell.style = index_style
        
    for row_idx, row in mix_xl_a.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=10 + row_idx, column=col_idx, value=val)
            if col_idx in [2, 3, 5]:
                cell.style = text_style
            elif col_idx in [6]:
                cell.style = integer_style
            else:
                cell.style = percent_style
                    
    ws.cell(row=12+mix_xl_a.shape[0], column=2, value="미디어믹스 B").style = index_style

    for col_idx, col_name in enumerate(mix_xl_b.columns, start=2):
        cell = ws.cell(row=13+mix_xl_a.shape[0], column=col_idx, value=colname_dict[col_name])
        cell.style = index_style

    for row_idx, row in mix_xl_b.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=14+mix_xl_a.shape[0] + row_idx, column=col_idx, value=val)
            if col_idx in [2, 3, 5]:
                cell.style = text_style
            elif col_idx in [6]:
                cell.style = integer_style
            else:
                cell.style = percent_style
                
    ## 스펙트럼
    spec_xl_p = pd.DataFrame(reportOptimize['table_spec']['reach_p'])
    spec_xl_n = pd.DataFrame(reportOptimize['table_spec']['reach_n'])
    spec_xl_s = pd.DataFrame(reportOptimize['table_spec']['reach_scaled'])

    for col in range(2, spec_xl_p.shape[1] + 2):
        cell = ws.cell(row=14+mix_xl_a.shape[0] + mix_xl_b.shape[0], column=col)
        cell.border = bottom_border
        
    ws.cell(row=16+mix_xl_a.shape[0] + mix_xl_b.shape[0], column=2, value="도달 지표 ↓").style = index_style
    ws.cell(row=17+mix_xl_a.shape[0] + mix_xl_b.shape[0], column=2, value="Reach(%)").style = text_style

    metrics = ['Reach(%)', 'Reach(n)', 'Reach (Scaled)']
    metric_validator = DataValidation(
        type="list",
        formula1='"' + ",".join(metrics) + '"', 
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True
    )
    ws.add_data_validation(metric_validator)
    metric_validator.add(ws.cell(row=17+mix_xl_a.shape[0] + mix_xl_b.shape[0], column=2).coordinate)

    spec_area_cols = ['MIX A', 'MIX B', 'GRPs', 'GRPs (가중치+)', 'A.F', 
                    'R1+', 'R2+', 'R3+', 'R4+', 'R5+', 'R6+', 'R7+', 'R8+', 'R9+', 'R10+']
    for col_idx, col_name in enumerate(spec_area_cols, start=2):
        cell = ws.cell(row=19+mix_xl_a.shape[0] + mix_xl_b.shape[0], column=col_idx, value=col_name)
        cell.style = index_style
        
    for row_idx, row in spec_xl_n.iterrows():
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=20+mix_xl_a.shape[0] + mix_xl_b.shape[0] + row_idx, column=col_idx, value=val)
            
            if col_idx in [2, 3]:
                cell.style = text_style
                
            elif col_idx in [4, 5, 6]:
                row_ = 20+mix_xl_a.shape[0] + mix_xl_b.shape[0] + row_idx
                cell = ws.cell(
                    row=row_, column=col_idx,
                    value=f'= SUMIFS(RAW_DATA!{get_column_letter(col_idx)}:{get_column_letter(col_idx)}, RAW_DATA!$B:$B, 분석결과!$B{row_}, RAW_DATA!$C:$C, 분석결과!$C{row_}, RAW_DATA!$Q:$Q, 분석결과!$B${17+mix_xl_a.shape[0] + mix_xl_b.shape[0]})'
                )
                cell.style = float_style
            else:
                row_ = 20+mix_xl_a.shape[0] + mix_xl_b.shape[0] + row_idx
                cell = ws.cell(
                    row=row_, column=col_idx,
                    value=f'= SUMIFS(RAW_DATA!{get_column_letter(col_idx)}:{get_column_letter(col_idx)}, RAW_DATA!$B:$B, 분석결과!$B{row_}, RAW_DATA!$C:$C, 분석결과!$C{row_}, RAW_DATA!$Q:$Q, 분석결과!$B${17+mix_xl_a.shape[0] + mix_xl_b.shape[0]})'
                )
                cell.style = numeric_style

    for letter_ in [get_column_letter(i) for i in range(2, 18)]:
        if letter_ in ['B', 'C']:
            ws.column_dimensions[letter_].width = 18
        elif letter_ in ['D', 'E', 'F']:
            ws.column_dimensions[letter_].width = 15
        else :
            ws.column_dimensions[letter_].width = 11

    # RAW DATA
    ws2 = wb.create_sheet("RAW_DATA")
    ws2.sheet_view.zoomScale = 90
    ws2.column_dimensions['A'].width = 3.86
    ws2.row_dimensions[1].height = 24
    ws2.sheet_view.showGridLines = False
    ws2.sheet_state = 'veryHidden'
    ws2.row_dimensions[1].hidden = True  
    ws2.row_dimensions[2].hidden = True 

    spec_xl_prw = spec_xl_p.assign(metric="Reach(%)")
    spec_xl_prw.columns = spec_area_cols + ['metric']
    spec_xl_nrw = spec_xl_n.assign(metric="Reach(n)")
    spec_xl_nrw.columns = spec_area_cols + ['metric']
    spec_xl_srw = spec_xl_s.assign(metric="Reach (Scaled)")
    spec_xl_srw.columns = spec_area_cols + ['metric']

    spec_raw = pd.concat([spec_xl_prw, spec_xl_nrw, spec_xl_srw], 
        ignore_index=True
    )

    for col_idx, col_name in enumerate(spec_raw.columns, start=2):
        cell = ws2.cell(row=3, column=col_idx, value=col_name)
        cell.style = index_style
        
    for row_idx, row in spec_raw.iterrows():
        for col_idx, val in enumerate(row, start=2):       
            cell = ws2.cell(row=4 + row_idx, column=col_idx, value=val) 
            if col_idx in [2, 3]:
                cell.style = text_style
            elif col_idx in [4, 5, 6]:
                cell.style = float_style
            else:
                cell.style = numeric_style
                
    for col_idx in range(4, 17):
        ws2.cell(
            row=1, column=col_idx, 
            value=f"= MAX(분석결과!${get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]}:{get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]+spec_xl_n.shape[0]})"
        )
        ws2.cell(
            row=2, column=col_idx, 
            value=f"= MIN(분석결과!${get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]}:{get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]+spec_xl_n.shape[0]})"
        )

    for col_idx in range(4, 17):
        color_rule_ = ColorScaleRule(
            start_type='formula', start_value=f"RAW_DATA!${get_column_letter(col_idx)}$2", start_color="FFFFFF",
            end_type='formula', end_value=f"RAW_DATA!${get_column_letter(col_idx)}$1", end_color="F8696B"
        )
        ws.conditional_formatting.add(
            f"{get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]}:{get_column_letter(col_idx)}{20+mix_xl_a.shape[0] + mix_xl_b.shape[0]+spec_xl_n.shape[0]}", color_rule_
        )
        
    return wb