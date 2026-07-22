from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import openpyxl as op
import numpy as np

def DapMixSample(ChannelVehicleMap, userGrade):
    
    if userGrade == "B":
        media_col, prod_col = "A", "B"
        
        ChannelVehicleMap = dict(sorted(ChannelVehicleMap.items()))
        max_row = np.max([len(v) for v in ChannelVehicleMap.values()])
        max_col = len(ChannelVehicleMap.keys())

        mix_wb = op.Workbook()
        mix_wb.remove(mix_wb.active)

        mix_main = mix_wb.create_sheet('main')
        mix_main.sheet_view.zoomScale = 95
        mix_main.sheet_view.showGridLines = False
        mix_main.freeze_panes = 'A2' 

        mix_list = mix_wb.create_sheet('list')
        mix_list.sheet_view.zoomScale = 90

        # 스타일 정의
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')  # 연회색
        header_border = Border(bottom=Side(style='medium'))

        cell_font = Font(size=10)

        cell_alignment = Alignment(horizontal='center', vertical='center')

        ### mix_list

        # 상품 리스트
        mix_list['A1'] = "상품"
        mix_list['A1'].font = header_font
        mix_list['A1'].fill = header_fill
        mix_list['A1'].border = header_border
        mix_list['A1'].alignment = cell_alignment

        for r in range(2, max_row + 2): 
            text_ = f'상품{str(r - 1)}'
            cell_ = mix_list.cell(column=1, row=r)
            cell_.value = text_
            cell_.font = header_font
            cell_.border = Border(right=Side(style='medium'))
            cell_.alignment  = cell_alignment

        for k, platform in enumerate(ChannelVehicleMap.keys(), start=2):
            header_ = mix_list.cell(column=k, row=1)
            header_.value = platform
            header_.font = header_font
            header_.fill = header_fill
            header_.border = header_border
            header_.alignment = cell_alignment

            for v, vehicle in enumerate(ChannelVehicleMap[platform], start=2):
                cell_ = mix_list.cell(column=k, row=v)
                cell_.value = vehicle
                cell_.font = cell_font
                cell_.alignment  = cell_alignment

        # 단가 & 성별
        bid_type_col = get_column_letter(max_col+2)
        mix_list[f"{bid_type_col}1"] = '단가유형'
        for ii, bt in enumerate(['CPM', 'CPC', 'CPV', 'CPRP', 'E.IMP 직접입력', 'E.GRP 직접입력'], start=2):
            cell_ = mix_list.cell(column=max_col+2, row=ii)
            cell_.value = bt
        mix_list.column_dimensions[bid_type_col].hidden = True

        gender_col = get_column_letter(max_col+3)
        mix_list[f"{gender_col}1"] = '성별'
        for ii, bt in enumerate(['P','M','F'], start=2):
            cell_ = mix_list.cell(column=max_col+3, row=ii)
            cell_.value = bt

        mix_list.column_dimensions[gender_col].hidden = True

        ### mix_main

        main_cols = {
            '매체': 15,
            '광고상품': 20,
            '성별': 7,
            'min': 7,
            'max': 7,
            'Impact 가중치': 15,
            '예산': 15,
            '단가 유형': 15,
            '단가': 15,
            '효율': 10,
            'E.IMP': 15,
            'E.GRP': 15
        }

        main_validator = {
            '매체': DataValidation(
                type="list", formula1=f"=list!B$1:{get_column_letter(max_col+1)}$1", 
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            ),
            '성별': DataValidation(
                type="list", formula1=f"=list!{gender_col}$2:{gender_col}$4", 
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            ),
            '단가 유형': DataValidation(
                type="list", formula1=f"=list!{bid_type_col}$2:{bid_type_col}$7",
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            )
        }

        main_formatter = {
            '매체': '@',
            '광고상품': '@',
            '성별': '@',
            'min': '0',
            'max': '0',
            'Impact 가중치': '0.00%',
            '예산': '#,##0',
            '단가 유형': '@',
            '단가': '#,##0',
            '효율': '0.00%',
            'E.IMP': '#,##0',
            'E.GRP': '#,##0.00'
        }

        for k, col_ in enumerate(main_cols.keys(), start=1):
            col_letter = get_column_letter(k)
            mix_main.column_dimensions[col_letter].width = main_cols[col_]

            header_ = mix_main.cell(column=k, row=1)
            header_.value = col_
            header_.font = header_font
            header_.fill = header_fill
            header_.border = Border(bottom=Side(style='medium'), right=Side(style='thin'), left=Side(style='thin'))
            header_.alignment = cell_alignment

            for v, _ in enumerate(range(100), start=2):
                cell_ = mix_main.cell(column=k, row=v)
                cell_.border = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
                cell_.value = None
                cell_.font = cell_font
                cell_.alignment  = cell_alignment
                cell_.number_format = main_formatter[col_]

            if main_validator.get(col_):
                col_validator= main_validator[col_]
                mix_main.add_data_validation(col_validator)
                col_validator.add(f'{col_letter}2:{col_letter}101')
            else:
                pass

        for col_idx, platform in enumerate(ChannelVehicleMap.keys(), start=2):
            col_letter = get_column_letter(col_idx)
            start_row = 2
            end_row = len(ChannelVehicleMap[platform]) + 1 
            range_formula = f"list!${col_letter}${start_row}:${col_letter}${end_row}"
            
            safe_name = f"ID_{col_idx - 1}" 
            
            defined_name = DefinedName(name=safe_name, attr_text=range_formula)
            mix_wb.defined_names.add(defined_name)

        header_range = f"list!$B$1:{get_column_letter(max_col + 1)}$1"
        for row in range(2, 102):
            formula = f'=INDIRECT("ID_"&MATCH({media_col}{row}, {header_range}, 0))'
            
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            mix_main.add_data_validation(dv)
            dv.add(f"{prod_col}{row}")
    
    # userGrade != "B"
    else:
        media_col, prod_col = "B", "C"
        
        ChannelVehicleMap = dict(sorted(ChannelVehicleMap.items()))
        max_row = np.max([len(v) for v in ChannelVehicleMap.values()])
        max_col = len(ChannelVehicleMap.keys())

        mix_wb = op.Workbook()
        mix_wb.remove(mix_wb.active)

        mix_main = mix_wb.create_sheet('main')
        mix_main.sheet_view.zoomScale = 95
        mix_main.sheet_view.showGridLines = False
        mix_main.freeze_panes = 'A2' 

        mix_list = mix_wb.create_sheet('list')
        mix_list.sheet_view.zoomScale = 90

        # 스타일 정의
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')  # 연회색
        header_border = Border(bottom=Side(style='medium'))

        cell_font = Font(size=10)

        cell_alignment = Alignment(horizontal='center', vertical='center')

        ### mix_list

        # 상품 리스트
        mix_list['A1'] = "상품"
        mix_list['A1'].font = header_font
        mix_list['A1'].fill = header_fill
        mix_list['A1'].border = header_border
        mix_list['A1'].alignment = cell_alignment

        for r in range(2, max_row + 2): 
            text_ = f'상품{str(r - 1)}'
            cell_ = mix_list.cell(column=1, row=r)
            cell_.value = text_
            cell_.font = header_font
            cell_.border = Border(right=Side(style='medium'))
            cell_.alignment  = cell_alignment

        for k, platform in enumerate(ChannelVehicleMap.keys(), start=2):
            header_ = mix_list.cell(column=k, row=1)
            header_.value = platform
            header_.font = header_font
            header_.fill = header_fill
            header_.border = header_border
            header_.alignment = cell_alignment

            for v, vehicle in enumerate(ChannelVehicleMap[platform], start=2):
                cell_ = mix_list.cell(column=k, row=v)
                cell_.value = vehicle
                cell_.font = cell_font
                cell_.alignment  = cell_alignment

        # 단가 & 성별
        bid_type_col = get_column_letter(max_col+2)
        mix_list[f"{bid_type_col}1"] = '단가유형'
        for ii, bt in enumerate(['CPM', 'CPC', 'CPV', 'CPRP', 'E.IMP 직접입력', 'E.GRP 직접입력', '기집행 분석 (IMP)', '기집행 분석 (GRP)'], start=2):
            cell_ = mix_list.cell(column=max_col+2, row=ii)
            cell_.value = bt
        mix_list.column_dimensions[bid_type_col].hidden = True

        gender_col = get_column_letter(max_col+3)
        mix_list[f"{gender_col}1"] = '성별'
        for ii, bt in enumerate(['P','M','F'], start=2):
            cell_ = mix_list.cell(column=max_col+3, row=ii)
            cell_.value = bt

        mix_list.column_dimensions[gender_col].hidden = True

        ### mix_main

        main_cols = {
            '캠페인': 20,
            '매체': 15,
            '광고상품': 20,
            '시작일': 15,
            '종료일': 15,
            '성별': 7,
            'min': 7,
            'max': 7,
            '타겟팅 모수': 15,
            'Impact 가중치': 15,
            '예산': 15,
            '단가 유형': 15,
            '단가': 15,
            '효율': 10,
            'IMP / GRP': 15,
            'REACH': 15
        }

        main_validator = {
            '매체': DataValidation(
                type="list", formula1=f"=list!B$1:{get_column_letter(max_col+1)}$1", 
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            ),
            '성별': DataValidation(
                type="list", formula1=f"=list!{gender_col}$2:{gender_col}$4", 
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            ),
            '단가 유형': DataValidation(
                type="list", formula1=f"=list!{bid_type_col}$2:{bid_type_col}$9",
                allow_blank=True, showErrorMessage=True, showInputMessage=True
            )
        }

        main_formatter = {
            '캠페인': '@',
            '매체': '@',
            '광고상품': '@',
            '시작일': 'YYYY-MM-DD',
            '종료일': 'YYYY-MM-DD',
            '성별': '@',
            'min': '0',
            'max': '0',
            '타겟팅 모수': '0',
            'Impact 가중치': '0.00%',
            '예산': '#,##0',
            '단가 유형': '@',
            '단가': '#,##0',
            '효율': '0.00%',
            'IMP / GRP': '#,##0.00',
            'REACH': '#,##0.00'
        }

        for k, col_ in enumerate(main_cols.keys(), start=1):
            col_letter = get_column_letter(k)
            mix_main.column_dimensions[col_letter].width = main_cols[col_]

            header_ = mix_main.cell(column=k, row=1)
            header_.value = col_
            header_.font = header_font
            header_.fill = header_fill
            header_.border = Border(bottom=Side(style='medium'), right=Side(style='thin'), left=Side(style='thin'))
            header_.alignment = cell_alignment

            for v, _ in enumerate(range(100), start=2):
                cell_ = mix_main.cell(column=k, row=v)
                cell_.border = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
                cell_.value = None
                cell_.font = cell_font
                cell_.alignment  = cell_alignment
                cell_.number_format = main_formatter[col_]

            if main_validator.get(col_):
                col_validator= main_validator[col_]
                mix_main.add_data_validation(col_validator)
                col_validator.add(f'{col_letter}2:{col_letter}101')
            else:
                pass

        for col_idx, platform in enumerate(ChannelVehicleMap.keys(), start=2):
            col_letter = get_column_letter(col_idx)
            start_row = 2
            end_row = len(ChannelVehicleMap[platform]) + 1 
            range_formula = f"list!${col_letter}${start_row}:${col_letter}${end_row}"
            
            safe_name = f"ID_{col_idx - 1}" 
            
            defined_name = DefinedName(name=safe_name, attr_text=range_formula)
            mix_wb.defined_names.add(defined_name)

        header_range = f"list!$B$1:{get_column_letter(max_col + 1)}$1"
        for row in range(2, 102):
            formula = f'=INDIRECT("ID_"&MATCH({media_col}{row}, {header_range}, 0))'
            
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            mix_main.add_data_validation(dv)
            dv.add(f"{prod_col}{row}")
            
    return mix_wb