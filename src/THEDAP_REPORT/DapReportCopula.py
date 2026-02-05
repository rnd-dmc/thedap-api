from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd

from THEDAP_REPORT.DapReportStyler import *

def makeReportCopula(result_json):
    wb = Workbook()

    # ### 미디어믹스
    ws = wb.active
    ws.title = "매체간 통합&중복"
    
    ws.sheet_view.zoomScale = 90
    ws.column_dimensions['A'].width = 3.86
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False
    
    option_analysis = result_json['option_analysis']
    copula_union = pd.DataFrame(result_json['copula_union']).assign(copula_type='매체간 통합')
    copula_inter = pd.DataFrame(result_json['copula_inter']).assign(copula_type='매체간 단독/중복')

    copula_df = pd.concat([copula_union, copula_inter], axis=0, ignore_index=True)[['copula_type', 'comb', 'reach']]
    sum_budget = result_json['sum_budget']

    ##
    ws.cell(row=2, column=2, value="분석일자").style = index_style
    ws.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

    ws.cell(row=3, column=2, value="총 예산").style = index_style
    ws.cell(row=3, column=3, value=sum_budget).style = title_style

    ws.cell(row=4, column=2, value="분석기준 타겟").style = index_style
    ws.cell(row=4, column=3, value=f"{option_analysis['input_gender']}{option_analysis['input_agemin']}{option_analysis['input_agemax']}").style = title_style

    ws.cell(row=5, column=2, value="타겟 모수").style = index_style
    ws.cell(row=5, column=3, value=result_json['population']).style = title_style

    ws.cell(row=6, column=2, value="데이터 기준").style = index_style
    ws.cell(row=6, column=3, value=option_analysis['maxdate']).style = title_style
    
    ##    
    colname_dict = {
        'copula_type':'유형', 'comb':'매체 조합', 'reach':'매체 조합 도달률'
    }
    
    for col_idx, col_name in enumerate(copula_df.columns, start=2):
        if col_name == 'comb':
            ws.merge_cells(start_row=8, start_column=col_idx, end_row=8, end_column=col_idx+1)
            header_cell = ws.cell(row=8, column=col_idx, value=colname_dict[col_name])
            header_cell.style = index_style
            ws.cell(row=8, column=col_idx+1).style = index_style  # 오른쪽 셀에도 동일 적용
        elif col_name == 'reach':
            header_cell = ws.cell(row=8, column=col_idx+1, value=colname_dict[col_name])
            header_cell.style = index_style
        else:
            header_cell = ws.cell(row=8, column=col_idx, value=colname_dict[col_name])
            header_cell.style = index_style
    
    ##
    for row_idx, row in copula_df.iterrows():
        row_num = 9 + row_idx

        cell_type = ws.cell(row=row_num, column=2, value=row['copula_type'])
        cell_type.style = text_style

        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
        cell_comb = ws.cell(row=row_num, column=3, value=row['comb'])
        cell_comb.style = text_style

        cell_right = ws.cell(row=row_num, column=4)
        cell_right.style = text_style

        cell_reach = ws.cell(row=row_num, column=5, value=row['reach'])
        cell_reach.style = percent_style


    
    for letter_ in [get_column_letter(i) for i in range(2, 18)]:
        if letter_ in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[letter_].width = 20
            
    ws.auto_filter.ref = f"B8:B{8+len(copula_df)}" 
    
    return wb