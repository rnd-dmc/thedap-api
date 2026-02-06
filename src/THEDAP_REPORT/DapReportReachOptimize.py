from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd

from THEDAP_REPORT.DapReportStyler import *

def DapReportReachOptimize(reportOption, reportOptimize, opt_type, target_pop):
    if (opt_type == 'reach_max'):
        budget_options = [float(k.replace(',','')) for k in result_json['table_opt_max'][0].keys()]
        freq_xl = pd.DataFrame(result_json['table_freq_max'])
        tbl_xl = pd.concat([pd.DataFrame(result_json['table_opt_max'][0][k]).assign(budget_sum = float(k.replace(',',''))) for k in (result_json['table_opt_max'][0].keys())], ignore_index=True)
        tbl_xl2 = tbl_xl.query(f'budget_sum == {budget_options[0]}')
        mix_xl = pd.DataFrame(result_json['opt_mix_max'])
        option_xl = result_json['option_max']
        
    else:
        budget_options = [float(k.replace(',','')) for k in result_json['table_opt_target'][0].keys()]
        freq_xl = pd.DataFrame(result_json['table_freq_target'])
        tbl_xl = pd.concat([pd.DataFrame(result_json['table_opt_target'][0][k]).assign(budget_sum = float(k.replace(',', ''))) for k in (result_json['table_opt_target'][0].keys())], ignore_index=True)
        tbl_xl2 = tbl_xl.query(f'budget_sum == {budget_options[0]}')
        mix_xl = pd.DataFrame(result_json['opt_mix_target'])
        option_xl = result_json['option_target']

    wb = Workbook()
    try:
        ### 미디어믹스
        ws = wb.active
        ws.title = "분석결과"

        ws.sheet_view.zoomScale = 90
        ws.column_dimensions['A'].width = 3.86
        ws.row_dimensions[1].height = 24
        ws.sheet_view.showGridLines = False

        maxbudget = option_xl['opt_maxbudget']*1_000_000 if opt_type == 'reach_max' else budget_options[0]
        ##
        ws.cell(row=2, column=2, value="분석일자").style = index_style
        ws.cell(row=2, column=3, value=(datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")).style = title_style

        ws.cell(row=3, column=2, value="총 예산").style = index_style
        ws.cell(row=3, column=3, value=maxbudget).style = title_style

        ws.cell(row=4, column=2, value="분석기준 타겟").style = index_style
        ws.cell(row=4, column=3, value=f"{option_xl['input_gender']}{option_xl['input_agemin']}{option_xl['input_agemax']}").style = title_style

        ws.cell(row=5, column=2, value="타겟 모수").style = index_style
        ws.cell(row=5, column=3, value=target_pop).style = title_style
        
        ws.cell(row=6, column=2, value="분석 모델 버전").style = index_style
        ws.cell(row=6, column=3, value=option_xl['maxdate']).style = title_style
        ##
        colname_dict = {
            'campaign':'캠페인', 'platform':'광고 매체', 'product':'광고 상품', 'date_start':'시작일', 'date_end':'종료일', 'gender':'성별',
            'min':'min', 'max':'max', 'retargeting':'리타겟팅 모수', 'impact':'Impact 가중치', 
            'budget':'예산', 'bid_type':'단가유형', 'bid_cost':'단가', 'bid_rate':'효율', 'imp':'IMP / GRP', 'reach':'Reach', 'min_rat':'최소 할당(%)'
        }

        ws.cell(row=8, column=2, value="분석조건").style = index_style

        for col_idx, col_name in enumerate(mix_xl.columns, start=2):
            cell = ws.cell(row=9, column=col_idx, value=colname_dict[col_name])
            cell.style = index_style

        for row_idx, row in mix_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell = ws.cell(row=10 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3]:
                    cell.style = text_style
                elif col_idx in [4, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7]:
                    cell.style = percent_style
                elif col_idx in [6]:
                    cell.style = integer_style
                else:
                    cell.style = numeric_style2
                    
        ### RAW DATA
        ws2 = wb.create_sheet("RAW_DATA")
        ws2.sheet_view.zoomScale = 90
        ws2.column_dimensions['A'].width = 3.86
        ws2.row_dimensions[1].height = 24
        ws2.sheet_view.showGridLines = False
        ws2.sheet_state = 'veryHidden'

        colname_dict2 = {
            'line':'구분', 'campaign':'캠페인', 'platform':'광고 매체', 'product':'광고 상품', 'gender':'성별', 'age_min':'Min', 'age_max':'Max', 
            'budget':'예산', 'date_start':'시작일', 'date_end':'종료일', 'alloc_rat':'예산 비율(%)', 'budget_sum':'budget_sum',
            'target_impression':'Target Impression', 'target_impression_weighted':'Target Impression (가중치+)', 'target_grps':'GRPs', 'target_grps_weighted':'GRPs (가중치+)',
            'target_reach_n':'Target Reach(n)', 'target_reach2_n':'Target Reach 2+(n)', 'target_reach3_n':'Target Reach 3+(n)', 'target_reach4_n':'Target Reach 4+(n)', 'target_reach5_n':'Target Reach 5+(n)', 
            'target_reach6_n':'Target Reach 6+(n)', 'target_reach7_n':'Target Reach 7+(n)', 'target_reach8_n':'Target Reach 8+(n)', 'target_reach9_n':'Target Reach 9+(n)', 'target_reach10_n':'Target Reach 10+(n)', 
            'target_reach_p':'Reach(%)', 'target_reach2_p':'Reach 2+(%)', 'target_reach3_p':'Reach 3+(%)', 'target_reach4_p':'Reach 4+(%)', 'target_reach5_p':'Reach 5+(%)',
            'target_reach6_p':'Reach 6+(%)', 'target_reach7_p':'Reach 7+(%)', 'target_reach8_p':'Reach 8+(%)', 'target_reach9_p':'Reach 9+(%)', 'target_reach10_p':'Reach 10+(%)', 
            'target_af':'Target A.F', 'target_af_day':'Target A.F (1 Day)', 'target_af_week':'Target A.F (Week)', 'target_af_30':'Target A.F (30 Days)', 'imp_interval':'Impression Interval'
        }

        for col_idx, col_name in enumerate(tbl_xl.columns, start=2):
            cell = ws2.cell(row=2, column=col_idx, value=colname_dict2[col_name])
            cell.style = index_style

        for row_idx, row in tbl_xl.iterrows():
            for col_idx, val in enumerate(row, start=2):
                cell =ws2.cell(row=3 + row_idx, column=col_idx, value=val)
                if col_idx in [2, 3, 5]:
                    cell.style = text_style
                elif col_idx in [4, 7, 8]:
                    cell.style = percent_style
                else:
                    cell.style = numeric_style2

        ###
        ws.cell(row=12+mix_xl.shape[0], column=2, value="분석결과").style = index_style

        ws.cell(row=13+mix_xl.shape[0], column=2, value="예산 구분 ↓").style = index_style
        ws.cell(row=14+mix_xl.shape[0], column=2, value=budget_options[0]).style = text_style

        budget_validator = DataValidation(
            type="list",
            formula1='"' + ",".join([str(b) for b in budget_options]) + '"',
            allow_blank=False,
            showInputMessage=True,
            showErrorMessage=True
        )
        ws.add_data_validation(budget_validator)
        budget_validator.add(ws.cell(row=14+mix_xl.shape[0], column=2).coordinate)

        ws.merge_cells(start_row=14+mix_xl.shape[0], start_column=2, end_row=13+mix_xl.shape[0]+tbl_xl2.shape[0], end_column=2)

        for col_idx, col_name in enumerate(tbl_xl2.drop('budget_sum', axis=1).columns, start=3):
            cell = ws.cell(row=13+mix_xl.shape[0], column=col_idx, value=colname_dict2[col_name])
            cell.style = index_style

        for row_idx, row in tbl_xl2.drop('budget_sum', axis=1).iterrows():
            for col_idx, val in enumerate(row, start=3):
                cell = ws.cell(row=14+mix_xl.shape[0] + row_idx, column=col_idx, value=val)
                if col_idx in [3, 4]:
                    cell.style = text_style
                elif col_idx in [5, 17, 16, 15, 14, 14, 13, 12, 11, 10, 9, 8]:
                    cell.style = percent_style
                    cell.value = f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx}, RAW_DATA!$C:$C, 분석결과!$D{14+mix_xl.shape[0]+row_idx})' if (row_idx != (tbl_xl2.shape[0]-1)) else \
                        f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx})'
                elif col_idx in [6]:
                    cell.style = integer_style
                    cell.value = f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx}, RAW_DATA!$C:$C, 분석결과!$D{14+mix_xl.shape[0]+row_idx})' if (row_idx != (tbl_xl2.shape[0]-1)) else \
                        f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx})'
                else:
                    cell.style = float_style
                    cell.value = f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx}, RAW_DATA!$C:$C, 분석결과!$D{14+mix_xl.shape[0]+row_idx})' if (row_idx != (tbl_xl2.shape[0]-1)) else \
                        f'=SUMIFS(RAW_DATA!{get_column_letter(col_idx-1)}:{get_column_letter(col_idx-1)}, RAW_DATA!$Q:$Q, 분석결과!$B${14+mix_xl.shape[0]}, RAW_DATA!$B:$B, 분석결과!$C{14+mix_xl.shape[0]+row_idx})'

        for letter_ in [get_column_letter(i) for i in range(2, 18)]:
            if letter_ in ['B', 'C', 'D']:
                ws.column_dimensions[letter_].width = 20
            elif letter_ in ['F']:
                ws.column_dimensions[letter_].width = 15
            else:
                ws.column_dimensions[letter_].width = 12
                
        return wb

    except Exception as e:
        return wb