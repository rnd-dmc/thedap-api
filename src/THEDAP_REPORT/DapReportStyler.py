from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle, numbers
import uuid

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def make_style(fg_color, font_color, bold, halign, valign, num_format=None):
    name = f"style_{fg_color}_{font_color}_{halign}_{valign}_{uuid.uuid4().hex[:6]}"
    
    return NamedStyle(
        name=name,
        font=Font(size=10, color=font_color, bold=bold),
        fill=PatternFill(start_color=fg_color, end_color=fg_color, fill_type="solid"),
        border=thin_border,
        alignment=Alignment(horizontal=halign, vertical=valign),
        number_format=num_format or "General"
    )
    
index_style = make_style("404040", "FFFFFF", True, "center", "center")
title_style = make_style("FFFFFF", "000000", True, "center", "center", "#,##0;-#,##0;-")
text_style = make_style("FFFFFF", "000000", False, "center", "center", "#,##0 ;-#,##0 ;- ")
integer_style = make_style("FFFFFF", "000000", False, "right", "center", "#,##0 ;-#,##0 ;- ")
float_style = make_style("FFFFFF", "000000", False, "right", "center", "#,##0.00 ;-#,##0.00 ;- ")
percent_style_round = make_style("FFFFFF", "000000", False, "right", "center", "0% ; -0% ;- ")
percent_style = make_style("FFFFFF", "000000", False, "right", "center", "0.00% ; -0.00% ;- ")
percent_style2 = make_style("FFFFFF", "000000", False, "center", "center", "0.00% ; -0.00% ;- ")
numeric_style = make_style("FFFFFF", "000000", False, "center", "center", "[=0]- ;[<1]0.00% ;#,##0 ")
numeric_style2 = make_style("FFFFFF", "000000", False, "center", "center", "[=0]- ;[<1]0.00% ;#,##0.00 ")

gray_center_bold = {
    "fill": PatternFill(start_color="404040", end_color="404040", fill_type="solid"),
    "font": Font(color="FFFFFF", size=10, bold=True),
    "alignment": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    ),
}

gray_center_bold_small = {
    "fill": PatternFill(start_color="404040", end_color="404040", fill_type="solid"),
    "font": Font(color="FFFFFF", size=10, bold=True),
    "alignment": Alignment(horizontal="center", vertical="center"),
    "border": gray_center_bold["border"],
}

white_center_normal = {
    "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "font": Font(color="000000", size=9),
    "alignment": Alignment(horizontal="center", vertical="center"),
    "border": gray_center_bold["border"],
}

bottom_border = Border(bottom=Side(style="thin", color="000000"))