from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import openpyxl as op
import numpy as np

from THEDAP_REPORT.DapReportStyler import *

def DapCustomSample():
    
    wb = op.Workbook()
    wb.remove(wb.active)
    
    mix_main = wb.create_sheet('main')
    mix_main.sheet_view.zoomScale = 95
    mix_main.sheet_view.showGridLines = False
    mix_main.freeze_panes = 'A2' 
    
    # 스타일 정의
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')  # 연회색
    header_border = Border(bottom=Side(style='medium'))
    
    cell_font = Font(size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    
    main_cols = {
        'GRPs': 15,
        'Reach (%)': 15
    }
            
    main_formatter = {
        'GRPs': '#,##0.00',
        'Reach (%)': '#,##0.00',
    }
    
    for k, col_ in enumerate(main_cols.keys(), start=1):
        col_letter = get_column_letter(k)
        mix_main.column_dimensions[col_letter].width = main_cols[col_]
    
        header_ = mix_main.cell(column=k, row=1)
        header_.value = col_
        header_.font = header_font
        header_.fill = header_fill
        header_.border = Border(bottom=Side(style='medium'), right=Side(style='thin'), left=Side(style='thin'))
        header_.alignment = cell_alignment

        for v, _ in enumerate(range(100), start=2):
            cell_ = mix_main.cell(column=k, row=v)
            cell_.border = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
            cell_.value = None
            cell_.font = cell_font
            cell_.alignment  = cell_alignment
            cell_.number_format = main_formatter[col_]
        
    return wb